# Routes - Database (Import/Export)
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app import db
from app.models import (
    Producer, Loading, LoadingAllocation, Receipt, 
    Cooperative, Exporter
)
import io
import csv
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import json
    from geojson import FeatureCollection, Feature, Point
    GEOJSON_AVAILABLE = True
except ImportError:
    GEOJSON_AVAILABLE = False

database_bp = Blueprint('database', __name__, url_prefix='/api/database')

# Import Producers
@database_bp.route('/import-producers', methods=['POST'])
@jwt_required()
def import_producers():
    """Importer les producteurs depuis un fichier CSV"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.endswith('.csv'):
        return jsonify({'error': 'Only CSV files allowed'}), 400
    
    coop = Cooperative.query.first()
    if not coop:
        return jsonify({'error': 'Cooperative not configured'}), 400
    
    try:
        stream = io.TextIOWrapper(file.stream, encoding='utf-8')
        reader = csv.DictReader(stream)
        imported = 0
        
        for row in reader:
            producer = Producer(
                cooperative_id=coop.id,
                name=row.get('name', ''),
                phone=row.get('phone'),
                email=row.get('email'),
                village=row.get('village'),
                latitude=float(row.get('latitude', 0)) if row.get('latitude') else None,
                longitude=float(row.get('longitude', 0)) if row.get('longitude') else None,
                estimated_production=float(row.get('estimated_production', 0)) if row.get('estimated_production') else 0,
                is_active=row.get('is_active', 'true').lower() == 'true'
            )
            db.session.add(producer)
            imported += 1
        
        db.session.commit()
        return jsonify({'message': f'{imported} producers imported', 'count': imported}), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# Export Producers
@database_bp.route('/export-producers', methods=['GET'])
@jwt_required()
def export_producers():
    """Exporter les producteurs en Excel"""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'Excel export not available'}), 503
    
    coop = Cooperative.query.first()
    producers = Producer.query.filter_by(cooperative_id=coop.id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Producteurs'
    
    # Headers
    headers = ['ID', 'Nom', 'Téléphone', 'Email', 'Village', 'Latitude', 'Longitude', 'Production Est.', 'Production Liv.', 'Production Rest.']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for producer in producers:
        ws.append([
            producer.id,
            producer.name,
            producer.phone,
            producer.email,
            producer.village,
            producer.latitude,
            producer.longitude,
            producer.estimated_production,
            producer.production_delivered,
            producer.get_remaining_production()
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'producteurs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

# Export Loadings
@database_bp.route('/export-loadings', methods=['GET'])
@jwt_required()
def export_loadings():
    """Exporter les chargements en Excel"""
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'Excel export not available'}), 503
    
    coop = Cooperative.query.first()
    loadings = Loading.query.filter_by(cooperative_id=coop.id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Chargements'
    
    # Headers
    headers = ['N° Chargement', 'Date', 'Exportateur', 'Projet', 'Véhicule', 'Remorque', 'Conducteur', 'Connaissement', 'Poids Déclaré', 'Sacs', 'Statut']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Data
    for loading in loadings:
        ws.append([
            loading.loading_number,
            loading.loading_date.strftime('%Y-%m-%d'),
            loading.exporter.name if loading.exporter else '',
            loading.project.name if loading.project else '',
            loading.vehicle_number,
            loading.trailer_number,
            loading.driver_name,
            loading.bill_of_lading,
            loading.declared_weight,
            loading.total_sacks,
            loading.status
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'chargements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

# Export as GeoJSON
@database_bp.route('/export-geojson', methods=['GET'])
@jwt_required()
def export_geojson():
    """Exporter les données en GeoJSON"""
    if not GEOJSON_AVAILABLE:
        return jsonify({'error': 'GeoJSON export not available'}), 503
    
    coop = Cooperative.query.first()
    producers = Producer.query.filter_by(cooperative_id=coop.id).all()
    
    features = []
    for producer in producers:
        if producer.latitude and producer.longitude:
            feature = Feature(
                geometry=Point((producer.longitude, producer.latitude)),
                properties={
                    'id': producer.id,
                    'name': producer.name,
                    'village': producer.village,
                    'estimated_production': producer.estimated_production,
                    'production_delivered': producer.production_delivered,
                    'remaining_production': producer.get_remaining_production()
                }
            )
            features.append(feature)
    
    feature_collection = FeatureCollection(features)
    
    output = io.BytesIO()
    output.write(json.dumps(feature_collection).encode('utf-8'))
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/geo+json',
        as_attachment=True,
        download_name=f'producteurs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.geojson'
    )
